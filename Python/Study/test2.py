from openpyxl import load_workbook
import csv, ipaddress, time

start = time.time()  # 시작 시간 저장

# 변수
policy_file = 'D:/02_Programming/Python/Study/policy.xlsx'
range_file = 'D:/02_Programming/Python/Study/range.csv'
result_file = 'D:/02_Programming/Python/Study/result.csv'
range_list = []

wb = load_workbook(policy_file)  # Work Book
ws = wb.get_sheet_by_name('Sheet1')  # Work Sheet

'''
추출 시 조건
1. A열에는 정책명, B열에는 아이피를 넣는다.
2. 헤더는 없앤다.
3. 한 셀에 모든 비교 대상 아이피를 넣는다.
4. 아이피에 공백은 없어야 한다.
'''

rule = ws['A']  # A열(정책명) 추출
rule_list = [rule[x].value for x in range(len(rule))]

ip = ws['C']  # C열(목적지) 추출
target_list = [ip[x].value for x in range(len(ip))]

# 정책파일 검증, 공백 있는지 확인해야 함
if len(rule_list) != len(target_list):
    print("정책과 아이피 갯수가 맞지 않음")
    exit()

conv_target_list = [i.replace(' ','') for i in target_list]

total_count = len(conv_target_list)

# 대역 불러오기
with open(range_file, mode='r') as f:
    for row in f:
        range_list.append(row.rstrip('\n')) # 개행 제거 후 리스트에 저장

with open(result_file, mode='w', newline='') as result: # CSV 파일에 결과 저장
    wr = csv.writer(result)
    for row in range(0,len(conv_target_list)):
        percent = str(round(row/total_count*100,2))
        ip_in_cell = conv_target_list[row].split(',')
        for single_ip in ip_in_cell:
            if '.' not in single_ip:
                print(rule_list[row]+" 정책의 아이피 확인 바람 - "+single_ip)
                exit()

            for range in range_list:
                match = ipaddress.ip_address(single_ip) in ipaddress.ip_network(range) # 하나씩 비교하기
                if match:   # 매치되면 결과 출력 및 저장
                    print('['+percent+'%] '+rule_list[row]+' - '+single_ip+' - '+range+' - True')
                    wr.writerow([rule_list[row],single_ip,range,"True"])
                    break   # 매치되면 다음 비교
                else:
                    print('['+percent+'%] '+rule_list[row]+' - '+single_ip+' - '+range+' - False')
                    #wr.writerow([ip,range,"False"])

print("time :", time.time() - start)  # 현재시각 - 시작시간 = 실행 시간